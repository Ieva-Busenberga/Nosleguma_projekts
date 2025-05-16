import subprocess
import openpyxl
import re

# palaiž nolasīšanas skriptus
def run_scripts():
    print("Running scripts")
    subprocess.run(["python", "euronics.py"])
    subprocess.run(["python", "rdveikals.py"])
    print("Scripts completed")

# nolasa datus no excel faila un tos sakārto
def read_and_sort_excel(file_name):
    data = {}
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            nosaukums = row[0].value
            cena = str(row[1].value).replace(',','.') if row[1].value else None
            if nosaukums and cena and cena != 'nav':
                Cena = float(cena)
                
                euronics_match = re.search(r'Apple\s*(iPhone\s*\d+\s*(?:mini|Pro\s*Max|Pro|Plus)?)(?:,\s*(\d+)\s*GB)', nosaukums, re.IGNORECASE)
                if euronics_match:
                    model = euronics_match.group(1).strip()
                    gb = int(euronics_match.group(2)) if euronics_match.group(2) else None
                    if model and gb is not None:
                        if (model, gb) not in data:
                            data[(model,gb)] = []
                        data[(model, gb)].append(cena)
                    continue

                rdveikals_match = re.search(r'(iPhone\s*\d+\s*(?:mini|Pro\s*Max|Pro|Plus)?)\s*(\d+)GB', nosaukums, re.IGNORECASE)
                if rdveikals_match:
                    model = rdveikals_match.group(1).strip()
                    gb = int(rdveikals_match.group(2)) if rdveikals_match.group(2) else None
                    if model and gb is not None:
                        if (model, gb) not in data:
                            data[(model,gb)] = []
                        data[(model, gb)].append(cena)
                    continue

    except FileNotFoundError:
        print(f"Error: File '{file_name}' not found.")
        return{}
    
    sorted_data = {}
    for key in sorted(data.keys()):
        sorted_data[key] = sorted(data[key])
    return sorted_data

# salīdzina cenas starp veikaliem
def price_comparison(euronics_data, rdveikals_data, new_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Modelis'
    sheet['B1'] = 'Atmiņa (GB)'
    sheet['C1'] = 'Cena Euronics'
    sheet['D1'] = 'Cena RD Electronics'
    sheet['E1'] = 'Lētākais veikals'

    euronics_cheaper = 0
    rdveikals_cheaper = 0
    row_num = 2

    all_models = sorted(list(set(euronics_data.keys()) | set(rdveikals_data.keys())))

    for model, gb in all_models:
        price_euronics = euronics_data.get((model, gb), [None])[0]
        price_rdveikals = rdveikals_data.get((model, gb), [None])[0]
        cheaper = "" 

        if price_euronics is not None and price_rdveikals is not None:
            if price_euronics < price_rdveikals:
                cheaper = "Euronics"
                euronics_cheaper += 1
            elif price_rdveikals < price_euronics:
                cheaper = "RD Electronics"
                rdveikals_cheaper += 1
            else:
                cheaper = "-"
        elif price_euronics is not None:
            cheaper = "-"
        elif price_rdveikals is not None:
            cheaper = "-"
        
        sheet.append([model, gb, price_euronics, price_rdveikals, cheaper])
        row_num += 1

    workbook.save(new_file)
    total_count  = euronics_cheaper + rdveikals_cheaper
    euronics_pr = (euronics_cheaper / total_count * 100) if total_count  > 0 else 0
    rdveikals_pr = (rdveikals_cheaper / total_count * 100) if total_count > 0 else 0

    print(f"\nComparison price summary saved.")
    print(f"Total comparison count {total_count}.")
    print(f"Euronics was cheaper {euronics_cheaper} times ({euronics_pr:.2f}%).")
    print(f"RD Electronics was cheaper {rdveikals_cheaper} times ({rdveikals_pr:.2f}%).")  

if __name__ == "__main__":
    run_scripts()
    euronics_data = read_and_sort_excel('euronics.xlsx')
    rdveikals_data = read_and_sort_excel('rdveikals.xlsx')
    price_comparison(euronics_data, rdveikals_data, 'salidzinajums.xlsx')